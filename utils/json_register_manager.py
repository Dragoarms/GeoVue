# utils/json_register_manager.py

"""
JSON Register Manager for data storage with Excel Power Query integration.

This module handles all data operations using JSON files as the primary storage,
with Excel files using Power Query for read-only data display. This approach:
- Eliminates Excel file locking issues
- Allows for faster, more reliable data operations  
- Enables users to add custom columns in Excel without affecting the source data
- Provides automatic Power Query setup for new Excel files

The JSON structure is designed to be Power Query friendly, with separate files
for compartments and original images data.

Author: George Symonds
Created: 2025
"""

import os
import sys
import json
import time
import shutil
import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Union, Any
import threading
import pandas as pd
import numpy as np
from openpyxl import workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


class JSONRegisterManager:
    """
    Manages register data using JSON files with Excel Power Query integration.
    
    This class provides thread-safe access to drilling register data stored in JSON,
    with automatic creation of Excel files configured with Power Query connections.
    """
    
    # File names
    COMPARTMENT_JSON = "compartment_register.json"
    ORIGINAL_JSON = "original_images_register.json"
    REVIEW_JSON = "compartment_reviews.json"
    EXCEL_FILE = "Chip_Tray_Register.xlsx"
    DATA_SUBFOLDER = "Register Data (Do not edit)"
    
    # Sheet names
    COMPARTMENT_SHEET = "Compartment Register"
    MANUAL_SHEET = "Manual Entries"
    ORIGINAL_SHEET = "Original Images Register"
    REVIEW_SHEET = "Compartment Reviews"
    
    

    @staticmethod
    def check_existing_files_static(base_path: str) -> Dict[str, bool]:
        """
        Static method to check which files exist without creating an instance.
        
        Args:
            base_path: Directory path to check
            
        Returns:
            Dictionary with file existence status
        """
        base = Path(base_path)
        data_path = base / JSONRegisterManager.DATA_SUBFOLDER
        
        return {
            'excel': (base / JSONRegisterManager.EXCEL_FILE).exists(),
            'compartment_json': (data_path / JSONRegisterManager.COMPARTMENT_JSON).exists(),
            'original_json': (data_path / JSONRegisterManager.ORIGINAL_JSON).exists(),
            'review_json': (data_path / JSONRegisterManager.REVIEW_JSON).exists(),
            'data_folder': data_path.exists() and any(data_path.iterdir()) if data_path.exists() else False
        }
    
    @staticmethod
    def has_existing_data_static(base_path: str) -> bool:
        """Static method to check if any register data exists without creating an instance."""
        existing = JSONRegisterManager.check_existing_files_static(base_path)
        return any(existing.values())
    
    @staticmethod
    def get_data_summary_static(base_path: str) -> Dict[str, Any]:
        """
        Static method to get summary of existing data without creating an instance.
        
        Args:
            base_path: Directory path to check
            
        Returns:
            Dictionary with data counts and info
        """
        base = Path(base_path)
        data_path = base / JSONRegisterManager.DATA_SUBFOLDER
        
        summary = {
            'compartment_count': 0,
            'original_count': 0,
            'review_count': 0,
            'has_excel': (base / JSONRegisterManager.EXCEL_FILE).exists(),
            'has_json_data': False
        }
        
        try:
            comp_path = data_path / JSONRegisterManager.COMPARTMENT_JSON
            if comp_path.exists():
                with open(comp_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    summary['compartment_count'] = len(data)
                    summary['has_json_data'] = True
                    
            orig_path = data_path / JSONRegisterManager.ORIGINAL_JSON
            if orig_path.exists():
                with open(orig_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    summary['original_count'] = len(data)
                    summary['has_json_data'] = True
                    
            review_path = data_path / JSONRegisterManager.REVIEW_JSON
            if review_path.exists():
                with open(review_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    summary['review_count'] = len(data)
                    summary['has_json_data'] = True
                    
        except Exception:
            pass  # Silent fail for static check
            
        return summary


    def __init__(self, base_path: str, logger: Optional[logging.Logger] = None):
        """
        Initialize the JSON Register Manager.
        
        Args:
            base_path: Directory path for register files
            logger: Optional logger instance
        """
        self.base_path = Path(base_path)
        self.logger = logger or logging.getLogger(__name__)
        
        # Excel file goes in the base path
        self.excel_path = self.base_path / self.EXCEL_FILE
        
        # Data files go in subfolder
        self.data_path = self.base_path / self.DATA_SUBFOLDER
        
        # File paths for JSON data
        self.compartment_json_path = self.data_path / self.COMPARTMENT_JSON
        self.original_json_path = self.data_path / self.ORIGINAL_JSON
        self.review_json_path = self.data_path / self.REVIEW_JSON
        
        # Lock files in data folder
        self.compartment_lock = self.compartment_json_path.with_suffix('.json.lock')
        self.original_lock = self.original_json_path.with_suffix('.json.lock')
        self.review_lock = self.review_json_path.with_suffix('.json.lock')

        # Thread lock
        self._thread_lock = threading.Lock()
        
        # Initialize files if needed
        self._initialize_files()
    

    def check_existing_files(self) -> Dict[str, bool]:
        """
        Check which files already exist.
        
        Returns:
            Dictionary with file existence status
        """
        return {
            'excel': self.excel_path.exists(),
            'compartment_json': self.compartment_json_path.exists(),
            'original_json': self.original_json_path.exists(),
            'review_json': self.review_json_path.exists(),
            'data_folder': self.data_path.exists() and any(self.data_path.iterdir())
        }
    
    def has_existing_data(self) -> bool:
        """Check if any register data already exists."""
        existing = self.check_existing_files()
        return any(existing.values())
    
    def get_data_summary(self) -> Dict[str, Any]:
        """
        Get summary of existing data.
        
        Returns:
            Dictionary with data counts and info
        """
        summary = {
            'compartment_count': 0,
            'original_count': 0,
            'review_count': 0,
            'has_excel': self.excel_path.exists(),
            'has_json_data': False
        }
        
        try:
            if self.compartment_json_path.exists():
                data = self._read_json_file(self.compartment_json_path, self.compartment_lock)
                summary['compartment_count'] = len(data)
                summary['has_json_data'] = True
                
            if self.original_json_path.exists():
                data = self._read_json_file(self.original_json_path, self.original_lock)
                summary['original_count'] = len(data)
                summary['has_json_data'] = True
                
            if self.review_json_path.exists():
                data = self._read_json_file(self.review_json_path, self.review_lock)
                summary['review_count'] = len(data)
                summary['has_json_data'] = True
                
        except Exception as e:
            self.logger.error(f"Error getting data summary: {e}")
            
        return summary

    def _initialize_files(self) -> None:
        """Initialize JSON files and Excel with Power Query if they don't exist."""
        # Create directory
        self.base_path.mkdir(parents=True, exist_ok=True)
        self.logger.info(f"Ensured base path exists: {self.base_path}")

        self.data_path.mkdir(parents=True, exist_ok=True)
        self.logger.info(f"Ensured data path exists: {self.data_path}")
        
        # Initialize JSON files only if they don't exist
        if not self.compartment_json_path.exists():
            self._write_json_file(self.compartment_json_path, self.compartment_lock, [])
            self.logger.info(f"Created compartment JSON: {self.compartment_json_path}")
        else:
            self.logger.info(f"Compartment JSON already exists: {self.compartment_json_path}")
        
        if not self.original_json_path.exists():
            self._write_json_file(self.original_json_path, self.original_lock, [])
            self.logger.info(f"Created original images JSON: {self.original_json_path}")
        else:
            self.logger.info(f"Original images JSON already exists: {self.original_json_path}")
        
        # Initialize review JSON file only if it doesn't exist
        if not self.review_json_path.exists():
            self._write_json_file(self.review_json_path, self.review_lock, [])
            self.logger.info(f"Created review JSON: {self.review_json_path}")
        else:
            self.logger.info(f"Review JSON already exists: {self.review_json_path}")
        

        # Use template instead of creating from scratch
        # Create Excel from template only if it doesn't exist
        if not self.excel_path.exists():
            self.logger.info(f"Excel file does not exist, creating from template: {self.excel_path}")
            self._create_excel_from_template()
        else:
            self.logger.info(f"Excel file already exists: {self.excel_path}")
    
    def _create_excel_from_template(self) -> None:
        """Properly create Excel file from a .xltx template."""
        try:
            # Determine template path
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

            template_path = os.path.join(base_path, "resources", "Register Template File", "Chip_Tray_Register.xltx")

            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found at: {template_path}")

            self.logger.info(f"Opening template file: {template_path}")

            # Open as a workbook and save as standard .xlsx
            wb = load_workbook(template_path)
            wb.save(self.excel_path)

            if self.excel_path.exists():
                self.logger.info(f"Excel file created successfully at: {self.excel_path}")
            else:
                raise FileNotFoundError(f"Excel file was not created at: {self.excel_path}")

        except Exception as e:
            self.logger.error(f"Error creating Excel from template: {e}")
            self.logger.error(traceback.format_exc())
            raise
    
    def _acquire_file_lock(self, lock_path: Path, timeout: int = 30) -> bool:
        """Acquire file lock with timeout."""
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            try:
                if lock_path.exists():
                    # Check if lock is stale (older than 60 seconds)
                    lock_age = time.time() - lock_path.stat().st_mtime
                    if lock_age > 60:
                        self.logger.warning(f"Removing stale lock: {lock_path}")
                        lock_path.unlink()
                    else:
                        time.sleep(0.1)
                        continue
                
                # Create lock
                lock_path.write_text(str(os.getpid()))
                return True
                
            except Exception:
                time.sleep(0.1)
        
        return False
    
    def _release_file_lock(self, lock_path: Path) -> None:
        """Release file lock."""
        try:
            if lock_path.exists():
                lock_path.unlink()
        except Exception as e:
            self.logger.warning(f"Error releasing lock {lock_path}: {e}")
    
    def _read_json_file(self, file_path: Path, lock_path: Path) -> List[Dict]:
        """Read JSON file with locking."""
        if not self._acquire_file_lock(lock_path):
            raise RuntimeError(f"Could not acquire lock for {file_path}")
        
        try:
            if file_path.exists():
                with open(file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return []
        finally:
            self._release_file_lock(lock_path)
    
    def _write_json_file(self, file_path: Path, lock_path: Path, data: List[Dict]) -> None:
        """Write JSON file with locking and backup."""
        if not self._acquire_file_lock(lock_path):
            raise RuntimeError(f"Could not acquire lock for {file_path}")
        
        try:
            # Create backup if file exists
            if file_path.exists():
                backup_path = file_path.with_suffix('.json.backup')
                shutil.copy2(file_path, backup_path)
            
            # Write data with proper formatting for readability
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
                
        finally:
            self._release_file_lock(lock_path)
    
    def update_compartment(self, hole_id: str, depth_from: int, depth_to: int,
                         photo_status: str, approved_by: Optional[str] = None,
                         comments: Optional[str] = None) -> bool:
        """Update or create a compartment entry."""
        with self._thread_lock:
            try:
                # Read current data
                data = self._read_json_file(self.compartment_json_path, self.compartment_lock)
                
                # Convert to records for easier manipulation
                records = {(r['HoleID'], r['From'], r['To']): r for r in data}
                
                # ensure depths are 
                depth_from = int(depth_from)
                depth_to = int(depth_to)

                # Update or create record
                key = (hole_id, depth_from, depth_to)
                timestamp = datetime.now().isoformat()
                username = approved_by or os.getenv("USERNAME", "Unknown")
                
                if key in records:
                    # Update existing
                    records[key].update({
                        'Photo_Status': photo_status,
                        'Approved_Date': timestamp,
                        'Approved_By': username
                    })
                    if comments:
                        records[key]['Comments'] = comments
                else:
                    # Create new
                    records[key] = {
                        'HoleID': hole_id,
                        'From': depth_from,
                        'To': depth_to,
                        'Photo_Status': photo_status,
                        'Approved_Date': timestamp,
                        'Approved_By': username,
                        'Comments': comments
                    }
                
                # Convert back to list and save
                data = list(records.values())
                self._write_json_file(self.compartment_json_path, self.compartment_lock, data)
                
                self.logger.info(f"Updated compartment: {hole_id} {depth_from}-{depth_to}")
                return True
                
            except Exception as e:
                self.logger.error(f"Error updating compartment: {e}")
                return False
    
    def update_original_image(self, hole_id: str, depth_from: int, depth_to: int,
                            original_filename: str, is_approved: bool,
                            upload_success: bool, uploaded_by: Optional[str] = None,
                            comments: Optional[str] = None, file_count: Optional[int] = None,
                            all_filenames: Optional[str] = None) -> bool:
        """Update or create an original image entry."""
        with self._thread_lock:
            try:
                # Read current data
                data = self._read_json_file(self.original_json_path, self.original_lock)
                
                # Ensure depths are integers
                depth_from = int(depth_from)
                depth_to = int(depth_to)
                
                # Convert to records
                records = {(r['HoleID'], r['Depth_From'], r['Depth_To']): r for r in data}
                
                # Update or create record
                key = (hole_id, depth_from, depth_to)
                timestamp = datetime.now().isoformat()
                username = uploaded_by or os.getenv("USERNAME", "Unknown")
                status = "Uploaded" if upload_success else "Failed"
                
                if key in records:
                    # Update existing
                    record = records[key]
                    record['Original_Filename'] = original_filename
                    record['Uploaded_By'] = username
                    
                    # NEW: Update file count and all filenames if provided
                    if file_count is not None:
                        record['File_Count'] = file_count
                    if all_filenames is not None:
                        record['All_Filenames'] = all_filenames
                    
                    if is_approved:
                        record['Approved_Upload_Date'] = timestamp
                        record['Approved_Upload_Status'] = status
                    else:
                        record['Rejected_Upload_Date'] = timestamp  
                        record['Rejected_Upload_Status'] = status
                    
                    if comments:
                        record['Comments'] = comments
                else:
                    # Create new
                    records[key] = {
                        'HoleID': hole_id,
                        'Depth_From': depth_from,
                        'Depth_To': depth_to,
                        'Original_Filename': original_filename,
                        'File_Count': file_count if file_count is not None else 1,
                        'All_Filenames': all_filenames,
                        'Approved_Upload_Date': timestamp if is_approved else None,
                        'Approved_Upload_Status': status if is_approved else None,
                        'Rejected_Upload_Date': timestamp if not is_approved else None,
                        'Rejected_Upload_Status': status if not is_approved else None,
                        'Uploaded_By': username,
                        'Comments': comments
                    }
                
                # Convert back to list and save
                data = list(records.values())
                self._write_json_file(self.original_json_path, self.original_lock, data)
                
                self.logger.info(f"Updated original image: {hole_id} {depth_from}-{depth_to} (files: {file_count or 1})")
                return True
                
            except Exception as e:
                self.logger.error(f"Error updating original image: {e}")
                return False
    
    def update_compartment_review(self, hole_id: str, depth_from: int, depth_to: int,
                                reviewed_by: Optional[str] = None, comments: Optional[str] = None,
                                review_number: Optional[int] = None,  # NEW PARAMETER
                                **kwargs) -> bool:
        """
        Update or create a user-specific compartment review entry.
        Each user has ONE row per compartment, with review_number tracking edits.
        
        This method preserves all existing fields in the JSON even if they're not 
        in the current config, ensuring no data loss when toggles are modified.
        
        Args:
            hole_id: Hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
            reviewed_by: Person who reviewed (defaults to current user)
            comments: Review comments
            review_number: Explicit review number (if None, auto-increments)
            **kwargs: Additional fields from config (toggle fields, etc.)
        """
        with self._thread_lock:
            try:
                # Read current data
                data = self._read_json_file(self.review_json_path, self.review_lock)
                
                # ENSURE DEPTHS ARE INTEGERS
                depth_from = int(depth_from)
                depth_to = int(depth_to)

                # Get username
                username = reviewed_by or os.getenv("USERNAME", "Unknown")
                
                # Create timestamp
                timestamp = datetime.now().isoformat()
                
                # Find existing review by this user for this compartment
                existing_index = None
                existing_review = None
                for idx, review in enumerate(data):
                    if (review['HoleID'] == hole_id and 
                        review['From'] == depth_from and 
                        review['To'] == depth_to and
                        review['Reviewed_By'] == username):
                        existing_index = idx
                        existing_review = review.copy()  # Keep a copy to preserve all fields
                        break
                
                if existing_index is not None:
                    # Update existing review - preserve all existing fields
                    # ===================================================
                    # MODIFIED: Use explicit review number if provided
                    # ===================================================
                    if review_number is not None:
                        use_review_number = review_number
                    else:
                        use_review_number = existing_review.get('Review_Number', 0) + 1
                    
                    # Start with existing data to preserve all fields
                    updated_review = existing_review.copy()
                    
                    # Update core fields
                    updated_review.update({
                        'Review_Number': use_review_number,
                        'Review_Date': timestamp,
                        'Comments': comments,
                    })
                    
                    # Update fields from kwargs (new toggle values)
                    updated_review.update(kwargs)
                    
                    # Replace the record
                    data[existing_index] = updated_review
                    
                    self.logger.info(f"Updated existing review (edit #{use_review_number}) by {username} for: {hole_id} {depth_from}-{depth_to}")
                else:
                    # Create new review record
                    # ===================================================
                    # MODIFIED: Use explicit review number if provided
                    # ===================================================
                    use_review_number = review_number if review_number is not None else 1
                    
                    new_record = {
                        'HoleID': hole_id,
                        'From': depth_from,
                        'To': depth_to,
                        'Reviewed_By': username,
                        'Review_Number': use_review_number,
                        'Review_Date': timestamp,
                        'Initial_Review_Date': timestamp,
                        'Comments': comments,
                    }
                    
                    # Add all kwargs (toggle fields from config)
                    new_record.update(kwargs)
                    
                    # Add the new record
                    data.append(new_record)
                    
                    self.logger.info(f"Created new review #{use_review_number} by {username} for: {hole_id} {depth_from}-{depth_to}")
                
                # Save updated data
                self._write_json_file(self.review_json_path, self.review_lock, data)
                return True
                
            except Exception as e:
                self.logger.error(f"Error updating compartment review: {e}")
                self.logger.error(traceback.format_exc())
                return False
    
    def batch_update_compartments(self, updates: List[Dict]) -> int:
        """
        Batch update multiple compartment entries.
        
        Args:
            updates: List of dictionaries with keys:
                - hole_id: Hole ID
                - depth_from: Starting depth
                - depth_to: Ending depth
                - photo_status: Photo status
                - approved_by: Optional approver name
                - comments: Optional comments
                
        Returns:
            Number of successful updates
        """
        successful_updates = 0
        
        with self._thread_lock:
            try:
                # Read current data once
                data = self._read_json_file(self.compartment_json_path, self.compartment_lock)
                
                # Convert to records for easier manipulation
                records = {(r['HoleID'], r['From'], r['To']): r for r in data}
                
                # Process all updates
                for update in updates:
                    try:
                        hole_id = update['hole_id']
                        depth_from = int(update['depth_from'])
                        depth_to = int(update['depth_to'])
                        photo_status = update['photo_status']
                        
                        key = (hole_id, depth_from, depth_to)
                        timestamp = datetime.now().isoformat()
                        username = update.get('approved_by', os.getenv("USERNAME", "Unknown"))
                        
                        if key in records:
                            # Update existing
                            records[key].update({
                                'Photo_Status': photo_status,
                                'Approved_Date': timestamp,
                                'Approved_By': username
                            })
                            if update.get('comments'):
                                records[key]['Comments'] = update['comments']
                        else:
                            # Create new
                            records[key] = {
                                'HoleID': hole_id,
                                'From': depth_from,
                                'To': depth_to,
                                'Photo_Status': photo_status,
                                'Approved_Date': timestamp,
                                'Approved_By': username,
                                'Comments': update.get('comments')
                            }
                        
                        successful_updates += 1
                        
                    except Exception as e:
                        self.logger.error(f"Error in batch update for {update}: {e}")
                
                # Write all updates at once
                if successful_updates > 0:
                    data = list(records.values())
                    self._write_json_file(self.compartment_json_path, self.compartment_lock, data)
                    self.logger.info(f"Batch updated {successful_updates} compartments")
                
            except Exception as e:
                self.logger.error(f"Error in batch_update_compartments: {e}")
                
        return successful_updates

    def batch_update_compartment_colors(self, color_updates: List[Dict]) -> int:
        """
        Batch update compartment hex colors.
        
        Args:
            color_updates: List of dictionaries with keys:
                - hole_id: Hole ID
                - depth_from: Starting depth
                - depth_to: Ending depth  
                - average_hex_color: Hex color string
                
        Returns:
            Number of successful updates
        """
        if not color_updates:
            return 0
            
        successful_updates = 0
        
        with self._thread_lock:
            try:
                # Read current data
                data = self._read_json_file(self.compartment_json_path, self.compartment_lock)
                
                # Convert to records for easier manipulation
                records = {(r['HoleID'], r['From'], r['To']): r for r in data}
                
                # Apply updates
                for update in color_updates:
                    try:
                        hole_id = update['hole_id']
                        depth_from = int(update['depth_from'])
                        depth_to = int(update['depth_to'])
                        hex_color = update['average_hex_color']
                        
                        key = (hole_id, depth_from, depth_to)
                        
                        if key in records:
                            # Add or update the Average_Hex_Color field
                            records[key]['Average_Hex_Color'] = hex_color
                            successful_updates += 1
                        else:
                            self.logger.warning(f"No matching compartment found for {hole_id} {depth_from}-{depth_to}m")
                    
                    except Exception as e:
                        self.logger.error(f"Error updating color for {update}: {e}")
                
                # Save if we made updates
                if successful_updates > 0:
                    data = list(records.values())
                    self._write_json_file(self.compartment_json_path, self.compartment_lock, data)
                    self.logger.info(f"Successfully updated {successful_updates} hex colors")
                
            except Exception as e:
                self.logger.error(f"Error batch updating hex colors: {e}")
                
        return successful_updates

    def get_user_review(self, hole_id: str, depth_from: int, depth_to: int, 
                       username: Optional[str] = None) -> Optional[Dict]:
        """
        Get the review by a specific user for a compartment.
        Returns ALL fields stored in the JSON, not just current config fields.
        
        Args:
            hole_id: Hole identifier
            depth_from: Starting depth
            depth_to: Ending depth
            username: User to get review for (defaults to current user)
            
        Returns:
            Review dictionary with all stored fields or None if no review exists
        """
        with self._thread_lock:
            try:
                data = self._read_json_file(self.review_json_path, self.review_lock)
                
                # Default to current user
                username = username or os.getenv("USERNAME", "Unknown")
                
                # Find review for this compartment by this user
                for review in data:
                    if (review['HoleID'] == hole_id and 
                        review['From'] == depth_from and 
                        review['To'] == depth_to and
                        review['Reviewed_By'] == username):
                        return review.copy()  # Return a copy to prevent accidental modification
                
                return None
                
            except Exception as e:
                self.logger.error(f"Error getting user review: {e}")
                return None
    
    def get_all_reviews_for_compartment(self, hole_id: str, depth_from: int, depth_to: int) -> List[Dict]:
        """
        Get all reviews for a specific compartment (one per user).
        Returns ALL fields for each review, preserving legacy data.
        
        Returns:
            List of review dictionaries with all stored fields
        """
        with self._thread_lock:
            try:
                data = self._read_json_file(self.review_json_path, self.review_lock)
                
                # Filter reviews for this compartment
                reviews = [r.copy() for r in data 
                          if r['HoleID'] == hole_id 
                          and r['From'] == depth_from 
                          and r['To'] == depth_to]
                
                # Sort by most recent review date
                reviews.sort(key=lambda x: x.get('Review_Date', ''), reverse=True)
                return reviews
                
            except Exception as e:
                self.logger.error(f"Error getting all reviews: {e}")
                return []

    def get_compartment_data(self, hole_id: Optional[str] = None) -> pd.DataFrame:
        """Get compartment data as DataFrame."""
        with self._thread_lock:
            try:
                data = self._read_json_file(self.compartment_json_path, self.compartment_lock)
                df = pd.DataFrame(data)

                # Ensure Average_Hex_Color column exists for backwards compatibility
                if not df.empty and 'Average_Hex_Color' not in df.columns:
                    df['Average_Hex_Color'] = None
                
                
                if hole_id and not df.empty:
                    df = df[df['HoleID'] == hole_id]
                
                return df
                
            except Exception as e:
                self.logger.error(f"Error getting compartment data: {e}")
                return pd.DataFrame()
    
    def get_original_image_data(self, hole_id: Optional[str] = None) -> pd.DataFrame:
        """Get original image data as DataFrame."""
        with self._thread_lock:
            try:
                data = self._read_json_file(self.original_json_path, self.original_lock)
                df = pd.DataFrame(data)
                
                if hole_id and not df.empty:
                    df = df[df['HoleID'] == hole_id]
                
                return df
                
            except Exception as e:
                self.logger.error(f"Error getting original image data: {e}")
                return pd.DataFrame()
    
    def get_review_data(self, hole_id: Optional[str] = None) -> pd.DataFrame:
        """
        Get review data as DataFrame.
        Returns ALL fields stored in the JSON, not just current config fields.
        """
        with self._thread_lock:
            try:
                data = self._read_json_file(self.review_json_path, self.review_lock)
                df = pd.DataFrame(data)
                
                if hole_id and not df.empty:
                    df = df[df['HoleID'] == hole_id]
                
                return df
                
            except Exception as e:
                self.logger.error(f"Error getting review data: {e}")
                return pd.DataFrame()


    def _check_missing_compartments(self) -> Dict:
        """
        Check for missing compartments based on original image ranges.
        
        Returns:
            Dictionary with results
        """
        results = {'missing': 0, 'errors': []}
        
        if not self.json_manager:
            return results
        
        # Get original images data
        originals_df = self.json_manager.get_original_image_data()
        if originals_df.empty:
            return results
        
        # Get compartment data
        compartments_df = self.json_manager.get_compartment_data()
        
        # Build set of existing compartments
        existing_compartments = set()
        if not compartments_df.empty:
            for _, row in compartments_df.iterrows():
                if pd.notna(row['To']):
                    hole_id = row['HoleID']
                    depth_to = int(row['To'])
                    existing_compartments.add((hole_id, depth_to))
        
        # Check each original image range
        interval = self.config.get('compartment_interval', 1)
        
        for _, orig_row in originals_df.iterrows():
            if pd.notna(orig_row['Depth_From']) and pd.notna(orig_row['Depth_To']):
                hole_id = orig_row['HoleID']
                depth_from = int(orig_row['Depth_From'])
                depth_to = int(orig_row['Depth_To'])
                
                # Calculate expected compartments
                current_depth = depth_from
                while current_depth < depth_to:
                    comp_depth_to = current_depth + interval
                    
                    # Check if this compartment exists
                    if (hole_id, comp_depth_to) not in existing_compartments:
                        # ===================================================
                        # FIXED: Add to batch instead of immediate update
                        # ===================================================
                        self.compartment_updates.append({
                            'hole_id': hole_id,
                            'depth_from': current_depth,
                            'depth_to': comp_depth_to,
                            'photo_status': 'MISSING',
                            'approved_by': 'System Check',
                            'comments': f'Missing compartment detected from original range {depth_from}-{depth_to}m'
                        })
                        
                        results['missing'] += 1
                        existing_compartments.add((hole_id, comp_depth_to))
                        self.logger.info(f"Queued missing compartment: {hole_id} {current_depth}-{comp_depth_to}m")
                    
                    current_depth = comp_depth_to
        
        return results

    def get_review_field_summary(self) -> Dict[str, int]:
        """
        Get a summary of all fields that appear in review data.
        Useful for understanding what legacy fields exist.
        
        Returns:
            Dictionary mapping field names to count of records containing that field
        """
        with self._thread_lock:
            try:
                data = self._read_json_file(self.review_json_path, self.review_lock)
                
                field_counts = {}
                for record in data:
                    for field in record:
                        field_counts[field] = field_counts.get(field, 0) + 1
                
                return field_counts
                
            except Exception as e:
                self.logger.error(f"Error getting review field summary: {e}")
                return {}